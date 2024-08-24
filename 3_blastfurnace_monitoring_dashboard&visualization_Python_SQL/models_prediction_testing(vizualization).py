# -*- coding: utf-8 -*-
"""
Created on Thu Aug 24 14:31:20 2023

@author: kadilnikov
"""
## imports
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import openpyxl
import datetime
import statistics as st
import psycopg2
import io
import sys
import argparse
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from psycopg2.extras import DictCursor, RealDictCursor

matplotlib.rcParams['font.size'] = 6.0
pd.set_option ('display.max_colwidth', None)
pd.set_option ('display.max_row', None)
pd.set_option ('display.max_column', None)

class DB_Query:
    def __init__(self, dbname='XXXX', user='XXXXXXXXX',
                 password='XXXXXXXXXXX', host='XXXXXXXX'):
        self.dbname = dbname
        self.user = user
        self.password = password
        self.host = host

    def query(self, q_str, vars=None):
        try:
            conn = psycopg2.connect(dbname=self.dbname, user=self.user, 
                        password=self.password, host=self.host)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute(q_str, vars)
            df = cursor.fetchall()
            df = pd.DataFrame(df)
            return df
            
        except(Exception, Error) as error:
            print("Ошибка при работе с PostgreSQL", error)

        finally:
            if conn:
                cursor.close()
                conn.close()
                

## datetime settings

parser = argparse.ArgumentParser()
parser.add_argument("-d", "--date", help="input the date as: DD.MM.YY, default is currant date")
parser.add_argument("-f", "--file", help="enter the file name as: *.xlsx, default is 'models_preds_monitoring'")
args = parser.parse_args()
if args.date:
    try:
        TS1 = datetime.datetime.strptime(args.date,
                                    "%d.%m.%y").replace(hour=14, minute=00,
                                                        second=0, microsecond=0)
    except:
        
        try:
            TS1 = datetime.datetime.strptime(args.date,
                                             "%d.%m.%Y").replace(hour=14, minute=00,
                                                        second=0, microsecond=0)
        except:
            print('Wrong format of date. Using current date')
            TS1 = datetime.datetime.today().replace(hour=14, minute=00,
                                                second=0, microsecond=0)
else:
    TS1 = datetime.datetime.today().replace(hour=14, minute=00,
                                            second=0, microsecond=0)
if args.file:
    EXCEL_FILE = str(args.file)
else:
    EXCEL_FILE = "models_preds_monitoring.xlsx"  

#TS1 = datetime.datetime.fromisoformat('2023-08-31 08:00:00')
TS2 = TS1 - datetime.timedelta(days=1, hours=6)
print(f'Attempting data for the {TS1}')
#print(TS2, TS1)

## ETL

# manual journal failures loading
j_fails = pd.read_excel('j_fails_all.xlsx')
mapping = pd.read_excel('fail_reason_map.xlsx')

db = DB_Query()

# loading melts & release data
melts_df = db.query('SELECT p_id, p_begin, p_end, p_s2_begin, p_s2_end\
                   FROM public.report1\
                   WHERE p_end >= %s AND p_begin < %s AND furnace_id = 12\
                   ORDER BY p_begin', (TS2, TS1))


# loading models predictions
models_preds = db.query('SELECT r2.datetime_gen, r2.p_id, r2.model_type,\
                   r2.fault_type_int, r2.additive_weight, r2.additive_type,\
                   r2.fault_type\
                   FROM public.report2 r2\
                   JOIN public.report1 r1\
                   ON r1.p_id = r2.p_id\
                   WHERE r1.p_end >= %s AND r1.p_begin < %s AND r1.furnace_id = 12\
                   ORDER BY r2.datetime_gen', (TS2, TS1))

#print(models_preds)
        
# loading failures data
failures_df = db.query('SELECT dt_begin, dt_end, event_type, score\
                   FROM public.reportstop rs\
                   WHERE rs.dt_end >= %s AND rs.dt_begin < %s\
                   ORDER BY rs.dt_begin', (TS2, TS1))
    

# loading active energy data
features = db.query("WITH mod_bts AS\
                   (SELECT timestamp, w_akt_val0, minstralblock_hoist_val0,\
                   auto_prm1_val0, auto_prm2_val0, auto_prm3_val0,\
                   minstralp62psn_on_val0, upk,\
                   minstralw_vg_val0, minstralw_el1_val0,\
                   minstralw_el2_val0, minstralw_el3_val0,\
                   CASE\
                   WHEN extract (hour from timestamp) between 8 and 19\
                   THEN concat(extract (day from timestamp), '_2')\
                   WHEN extract (hour from timestamp) between 0 and 7\
                   THEN concat(extract (day from timestamp), '_1')\
                   WHEN extract (hour from timestamp) between 20 and 23\
                   THEN concat(extract (day from timestamp + interval '1 day'), '_1')\
                   END as day_rotation\
                   FROM public.bts\
                   WHERE timestamp >= %s AND timestamp < %s)\
                   SELECT timestamp, w_akt_val0, minstralblock_hoist_val0,\
                   auto_prm1_val0, auto_prm2_val0, auto_prm3_val0,\
                   minstralp62psn_on_val0, upk,\
                   minstralw_vg_val0, minstralw_el1_val0,\
                   minstralw_el2_val0, minstralw_el3_val0,\
                   sum(minstralw_el1_val0)\
                   over(partition by day_rotation order by timestamp asc) as w_el1_cumsum,\
                   sum(minstralw_el2_val0)\
                   over(partition by day_rotation order by timestamp asc) as w_el2_cumsum,\
                   sum(minstralw_el3_val0)\
                   over(partition by day_rotation order by timestamp asc) as w_el3_cumsum\
                   from mod_bts\
                   ORDER BY timestamp", (TS2, TS1))

   
# loading chemistry analytics
chemistry = db.query('SELECT *\
                   FROM public.probe_class pr\
                   WHERE rotation_begin >= %s AND rotation_begin < %s\
                   ORDER BY rotation_begin', (TS2, TS1))

chemistry = chemistry.round(3)


if models_preds.empty:
    sys.exit('No entries with models predictions in this period!\nCheck "report2" table')

#print(chemistry)


## data preprocessing
melts_df['m_duration'] = melts_df['p_end'] - melts_df['p_begin']
melts_df['duration_mins'] = melts_df['m_duration'].dt.seconds / 60

melts_df['r_duration'] = melts_df['p_s2_end'] - melts_df['p_s2_begin']
melts_df['r_duration_mins'] = melts_df['r_duration'].dt.seconds / 60


melts_df = melts_df.merge(models_preds.pivot_table(index='p_id',
                                                   columns='model_type',
                                                   values='fault_type_int',
                                                   dropna=False),
                          left_on='p_id', right_index=True, how='left')

melts_df = melts_df.merge(models_preds.pivot_table(index='p_id',
                                                   columns='model_type',
                                                   values='additive_weight',
                                                   dropna=False),
                          left_on='p_id', right_index=True, how='left',
                          suffixes=(None, '_weight'))

melts_df = melts_df.merge(models_preds.pivot_table(index='p_id',
                                                   columns='model_type',
                                                   values='additive_type',
                                                   dropna=False,
                                                   aggfunc='first'),
                          left_on='p_id', right_index=True, how='left',
                          suffixes=(None, '_addtype'))

failures_df['f_duration'] = failures_df['dt_end'] - failures_df['dt_begin']
failures_df['f_mins'] = failures_df['f_duration'].dt.seconds / 60

j_fails = j_fails[(j_fails['Дата и время окончания отклонения/неисправности'] > TS2)\
                   & (j_fails['Дата и время начала отклонения/неисправности'] < TS1)]

j_fails['duration'] = j_fails['Дата и время окончания отклонения/неисправности']\
                      - j_fails['Дата и время начала отклонения/неисправности']
#print(melts_df)
#print(failures_df)

def desc_opt(desc):
    result = desc.split(' \\ ')[-1]
    result = result[:-4]
    try:
        result = f'{mapping[mapping["Name"] == result]["group_num"].iloc[0]} - {result}'
    except:
        result = f'99 - {result}'
    return result


j_fails['optimized_desc'] = j_fails['Причины отклонения/неисправности'].apply(desc_opt)
j_fails = j_fails.sort_values('optimized_desc')

def add_chemflag(ts):
    chem_df = chemistry.loc[(chemistry['rotation_begin'] <= ts)\
              & (chemistry['rotation_end'] >= ts), 'total_class']
    #print(ts, chem_df, type(chem_df))
    try:
        result = chem_df.iloc[0]
    except:
        result = np.nan
    return result

if not chemistry.empty:
    
    features['chem_totalclass'] = features['timestamp'].apply(add_chemflag)

features['total'] = (features['auto_prm1_val0'] == 0) |\
                    (features['auto_prm2_val0'] == 0) |\
                    (features['auto_prm3_val0'] == 0) |\
                    (features['minstralp62psn_on_val0'] == 0) |\
                    (features['upk'] == 0) |\
                    (features['minstralblock_hoist_val0'] == 1)
                   
#print(melts_df)
#print(models_preds)
#print(failures_df)


melts_begins = melts_df['p_begin']
melts_ends = melts_df['p_end']
melts_bodies = melts_df[['p_begin', 'm_duration']].itertuples(index=False,
                                                              name=None)

releases_bodies = melts_df[['p_s2_begin', 'r_duration']].itertuples(index=False,
                                                              name=None)

#%%
## plot

LINES_LENGTH = 0.08
FAILURES_HEIGHT = 0.07
V_BORDERS = (-1.4, 0.5)
X_PERDAY = 20
Y = 9

fig, axs = plt.subplots(figsize=((TS1 - TS2).days * X_PERDAY, Y),
                       layout='constrained', nrows=6, ncols=1,
                       height_ratios=(0.28, 0.12, 0.2, 0.15, 0.1, 0.15), sharex=True)

ax = axs[0]

maj_locator = matplotlib.dates.HourLocator(interval=1)

min_locator = matplotlib.ticker.AutoMinorLocator(n=6)
min_locator.MAXTICKS = 10000

# melts and disorders plot
             
ax.broken_barh(list(melts_bodies), (V_BORDERS[0], V_BORDERS[1]-V_BORDERS[0]),
               facecolor='lavender')

ax.eventplot(melts_begins, colors='black', lineoffsets=0, 
                       linelengths=abs(V_BORDERS[0]) * 2,
             linewidth=0.3, alpha=1.0, label='Melts')

ax.eventplot(melts_ends, colors='black', lineoffsets=0,
             linelengths=abs(V_BORDERS[0]) * 2,
             linewidth=0.3, alpha=1.0)

disorders_list = {'1 electrod': [0.05, 'goldenrod'],
                  '3 electrod': [0.15, 'C0'],
                  'DT': [0.25, 'C1'],
                  'LR': [0.35, 'mediumseagreen'],
                  'best': [0.45, 'darkviolet']}

#depth_dict = {'Ckf,s


for index, row in melts_df.iterrows():
    for column in disorders_list:
        try:
            if row[column] == -1:
                ax.broken_barh([(row['p_begin'], row['m_duration'])],
                               (-disorders_list[column][0]-LINES_LENGTH/2, LINES_LENGTH),
                               facecolor=disorders_list[column][1])
                
                if column == 'best':
                    
                    label = models_preds[(models_preds['p_id'] == row['p_id'])\
                                         & (models_preds['model_type'] == column)]['fault_type']
                    
                    ax.text(row['p_begin'], -disorders_list[column][0]-LINES_LENGTH/2,
                    label.to_string(index=False), ha='left', va='bottom',
                    fontsize=5, color='white')
            
            if row[column] == 1:
                ax.broken_barh([(row['p_begin'], row['m_duration'])],
                               (disorders_list[column][0]-LINES_LENGTH/2, LINES_LENGTH),
                               facecolor=disorders_list[column][1])
                
                if column == 'best':
                    
                    label = models_preds.loc[(models_preds['p_id'] == row['p_id'])\
                                         & (models_preds['model_type'] == column), 'fault_type']
                    
                    ax.text(row['p_begin'], disorders_list[column][0]-LINES_LENGTH/2,
                    label.to_string(index=False), ha='left', va='bottom',
                    fontsize=5, color='white')

            if np.isnan(row[column]):
                ax.broken_barh([(row['p_begin'], row['m_duration'])],
                               (-disorders_list[column][0]-LINES_LENGTH/2, LINES_LENGTH),
                               facecolor='black')
                ax.broken_barh([(row['p_begin'], row['m_duration'])],
                               (disorders_list[column][0]-LINES_LENGTH/2, LINES_LENGTH),
                               facecolor='black')
        except:
            print(f'Column "{column}" is out in the table')
            
    ax.text(row['p_begin'] + (row['m_duration'] / 2),
            V_BORDERS[0] + 0.1,
            f'№{row["p_id"]:.0f}\n{row["duration_mins"]:.0f} min',
            ha='center', va='center', fontsize=7)

# failures plot
failures_list = {'1.Простой по мощности': [-0.61, 1],
                 '2.Облом электрода': [-0.7, 2],
                 '3.Отжег электрода': [-0.79, 3],
                 '4.Асимметрия': [-0.88, 4],
                 '5.Сбой работы газодувок': [-0.97, 5],
                 '6.Течь': [-1.06, 6],
                 '7.Неготовность инфраструктуры': [-1.15, 7]}

for fail_type in failures_list:
    failures = list(failures_df[failures_df['event_type'] == failures_list[fail_type][1]][['dt_begin',
                                                            'f_duration']].itertuples(index=False,
                                                                                  name=None))
    ax.broken_barh(failures,
                   (failures_list[fail_type][0], FAILURES_HEIGHT),
                   facecolor='lightgrey', edgecolor='darkgrey', linewidth=0.5)
    
    for fail in failures:

        #print(fail)
        if failures_list[fail_type][1] == 4:
            el_num = failures_df.loc[failures_df['dt_begin']==fail[0], 'score']
            ax.text(fail[0] + fail[1] / 2, failures_list[fail_type][0] + 0.03,
                f'{failures_list[fail_type][1]}_{int(el_num.iloc[0])}', ha='center', va='center',
                fontsize=6)
            
        else:
            ax.text(fail[0] + fail[1] / 2, failures_list[fail_type][0] + 0.03,
                    failures_list[fail_type][1], ha='center', va='center',
                    fontsize=6)

ax.tick_params("x", labelbottom=True)

# chemistry analytics


if not chemistry.empty:
    axs[1].plot(features['timestamp'], features['chem_totalclass'])
    axs[1].set_ylim(-2.1, 2.1)
    axs[1].yaxis.grid(visible=True, linewidth=0.2)
    axs[1].set_ylabel('chemistry', fontsize=8)
    for i, ser in chemistry.iterrows():
        #print(type(ser))
        #print(ser)
        axs[1].text(ser['rotation_begin']+datetime.timedelta(minutes=10), 0,
                    ser[5:9].to_string(), ha='left', va='center',
                    fontsize=6)
        
        axs[1].text(ser['rotation_begin']+datetime.timedelta(hours=2), 0,
                    ser[9:15].to_string(), ha='left', va='center',
                    fontsize=6)

        axs[1].text(ser['rotation_begin']+datetime.timedelta(hours=4), 0,
                    ser[15:21].to_string(), ha='left', va='center',
                    fontsize=6)

        axs[1].text(ser['rotation_begin']+datetime.timedelta(hours=6), 0,
                    ser[21:27].to_string(), ha='left', va='center',
                    fontsize=6)

        axs[1].text(ser['rotation_begin']+datetime.timedelta(hours=8), 0,
                    ser[27:33].to_string(), ha='left', va='center',
                    fontsize=6)
        
else:
    axs[1].text(TS2+datetime.timedelta(hours=2), 0,
                    'NO CHEMISTRY DATA IN THE DATABASE', ha='left', va='center',
                    fontsize=12)
    axs[1].set_ylim(-2.1, 2.1)
    axs[1].set_ylabel('chemistry', fontsize=8)

# act_energy plot

w_act_max = features['w_akt_val0'].max()

axs[2].eventplot(features[features['total'] == 1]['timestamp'],
                 colors='lightgrey',
                     lineoffsets=15, 
                     linelengths=30, linewidth=1, alpha=0.5) 

axs[2].plot(features['timestamp'], features['w_akt_val0'], label='w_akt_val0')
axs[2].plot(features['timestamp'], features['minstralw_el1_val0'], label='w_el1')
axs[2].plot(features['timestamp'], features['minstralw_el2_val0'], label='w_el2')
axs[2].plot(features['timestamp'], features['minstralw_el3_val0'], label='w_el3')

axs[2].plot(features['timestamp'], features['minstralw_vg_val0'], label='minstralw_vg_val0',
            color='orange', linewidth=0.5)

axs[2].broken_barh(list(releases_bodies), (0, w_act_max),
               facecolor='lightcyan', edgecolor='darkcyan')

axs[2].set_ylim(0, w_act_max)
axs[2].set_ylabel('w_akt', fontsize=8)
axs[2].yaxis.grid(visible=True, linewidth=0.2)
axs[2].legend()
#axs[2].tick_params("x", labelbottom=True)

# cumul energy
axs[3].plot(features['timestamp'], features['w_el1_cumsum'], label='w_el1_cumsum')
axs[3].plot(features['timestamp'], features['w_el2_cumsum'], label='w_el2_cumsum')
axs[3].plot(features['timestamp'], features['w_el3_cumsum'], label='w_el3_cumsum')
axs[3].yaxis.grid(visible=True, linewidth=0.2)
axs[3].set_ylabel('cumul_w_akt', fontsize=8)
axs[3].legend()

# flags plot

flags = ['auto_prm1_val0',
         'auto_prm2_val0', 'auto_prm3_val0',
         'minstralp62psn_on_val0', 'upk']

axs[4].eventplot(features[features['minstralblock_hoist_val0'] == 1]['timestamp'],
                 colors='brown',
                     lineoffsets=0, 
                     linelengths=1, linewidth=1, alpha=1.0), 
for flag in flags:
    axs[4].eventplot(features[features[flag] == 0]['timestamp'],
                     colors='brown',
                     lineoffsets=flags.index(flag) + 1, 
                     linelengths=1, linewidth=1, alpha=1.0)
                 
axs[4].set_yticks(range(len(flags) + 1), labels=['minstralblock_hoist_val0'] + flags)

axs[4].eventplot(melts_begins, colors='black', lineoffsets=len(flags)/2, 
                 linelengths=len(flags)+2,
                 linewidth=0.3, alpha=1.0)

axs[4].eventplot(melts_ends, colors='black', lineoffsets=len(flags)/2,
                 linelengths=len(flags)+2,
                 linewidth=0.3, alpha=1.0)


axs[4].set_ylim(-0.5, len(flags) + 0.5)
axs[4].tick_params("x", labelbottom=True)
axs[4].set_ylabel('control flags', fontsize=8)

# journal failures plot

j_fails_list = j_fails['optimized_desc'].unique()
#print(j_fails_list)

y_ticks = []
for fail_type in j_fails_list:
    
    j_fail = list(j_fails[j_fails['optimized_desc'] == fail_type]\
                  [['Дата и время начала отклонения/неисправности',\
                  'duration']].itertuples(index=False, name=None))

    pos_y = len(j_fails_list) - list(j_fails_list).index(fail_type)
    y_ticks.append(pos_y)
    axs[5].broken_barh(j_fail,
                       (pos_y - 0.5, 1),
                   facecolor=f'C{pos_y}', edgecolor='darkgrey', linewidth=0.5)
    
axs[5].set_yticks(y_ticks, labels=j_fails_list)
axs[5].yaxis.grid(visible=True, linestyle='dotted', linewidth=0.5)

axs[5].eventplot(melts_begins, colors='black', lineoffsets=len(j_fails_list)/2, 
                 linelengths=len(j_fails_list)+2,
                 linewidth=0.3, alpha=1.0)

axs[5].eventplot(melts_ends, colors='black', lineoffsets=len(j_fails_list)/2,
                 linelengths=len(j_fails_list)+2,
                 linewidth=0.3, alpha=1.0)

axs[5].set_ylim(0, len(j_fails_list) + 1)
axs[5].set_ylabel('journal_failures', fontsize=8)


# shift begin
ax.vlines(pd.date_range(TS2.replace(hour=8, minute=0, second=0, microsecond=0),
                         TS1, freq='12H'),
          *V_BORDERS, colors='purple', linestyle='--')

# other settings           
ax.set_ylabel('отказы                       \nнедостаток восстановителя                       \nизбыток восстановителя',
              fontsize=7, ha='left', va='bottom')

ax.set_yticks([x[0] for x in list(disorders_list.values())] +
              [-x[0] for x in list(disorders_list.values())] +
              [x[0]+0.03 for x in list(failures_list.values())],
              labels=2 * list(disorders_list.keys()) + list(failures_list.keys()),
              minor=False, fontdict={'fontsize': 7.5})

ax.xaxis.set_major_locator(maj_locator)
ax.xaxis.set_minor_locator(min_locator)
#ax.grid(visible=True, which='major', axis='x', color='grey', linestyle='-', linewidth=0.1)
#ax.grid(visible=True, which='minor', color='grey', linestyle='-', linewidth=0.05)
ax.set_xlim(min(melts_df.loc[0, 'p_begin'], TS2), max(melts_df.loc[len(melts_df)-1, 'p_end'], TS1)) # pd.Timestamp('2023-03-31 13:40:00'
ax.set_ylim(*V_BORDERS)

ax.hlines([0, 0.02 + FAILURES_HEIGHT + max([x[0] for x in failures_list.values()]),
           -0.02 + min([x[0] for x in failures_list.values()])],
          TS2, TS1, colors='black')
ax.set_title(f'Models predictions monitoring   |  {TS1}', loc='left')


## export to excel



try:
    wb = load_workbook(filename=EXCEL_FILE)
except:
    wb = openpyxl.Workbook()
    
sheet_name = f'{TS1.date()}'
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    wb.create_sheet(title=sheet_name, index=1)
    ws = wb[sheet_name]

# transform plot to image and add to excel
buf = io.BytesIO()
fig.savefig(buf)
buf.seek(0)
plot_img = Image(buf)
ws.add_image(plot_img, 'A1')

# write data to the table
if 'table' in wb.sheetnames:
    ws = wb['table']
else:
    wb.create_sheet(title='table')
    ws = wb['table']


colwr_list = {'p_begin': 'A', 'p_id': 'C',
              '1 electrod': 'D', '1 electrod_addtype': 'E', '1 electrod_weight': 'F',
              '3 electrod': 'G', '3 electrod_addtype': 'H', '3 electrod_weight': 'I',
              'DT': 'J', 'DT_addtype': 'K', 'DT_weight': 'L',
              'LR': 'M', 'LR_addtype': 'N', 'LR_weight': 'O',
              'best': 'P', 'best_addtype': 'Q', 'best_weight': 'R'}

def fill_data(row_in, row_out):
    for col in colwr_list:
        try:
            ws[f'{colwr_list[col]}{row_out}'] = melts_df.loc[row_in, col]
        except:
            continue
     
j = 0
for p_id in melts_df.loc[:, 'p_id']:
    i = 6
    for row in ws.iter_rows(min_row=6, min_col=3, max_col=3, max_row=1000, values_only=True):
        #print(f'p_id: {p_id}, row: {row[0]}')
        if p_id == row[0] or row[0] == None:
            fill_data(j, i)
            break
        
        elif p_id < row[0]:
            ws.insert_rows(i)
            fill_data(j, i)
            break
        i += 1
    j += 1

wb.save(EXCEL_FILE)

print(f'Plot added successfully.\n Plot path: {EXCEL_FILE}')
#fig.savefig(f'ep_{str(TS1.strftime("%d.%m.%y_%H.%M"))}.png', dpi=100)

#plt.show()
#%%

