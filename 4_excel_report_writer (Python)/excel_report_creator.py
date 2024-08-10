### component version
ver = 3.4

import os
import sys
import re

import numpy as np, pandas as pd
#from builtin_data import InputTable, InputTables, InputVariables, OutputTable, DataType, DataKind, UsageType
#from builtin_pandas_utils import to_data_frame, prepare_compatible_table, fill_table
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

import openpyxl
import string
import random
import datetime
import json
import logging
#logging.basicConfig(level=logging.DEBUG, filename="excelreport_log.log", filemode="w", format="%(relativeCreated)d %(lineno)d %(asctime)s %(levelname)s %(message)s") #"a" - append
from copy import copy

random.seed(8140)

pd.set_option('display.max_column', 300)

class ReportCreator:

    ### component version
    ver = 3.4
    
    # for collecting errors and history operations
    errors = {}
    writing_history = {}
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _path_check(self, income_path: str, filename: str, file_ext: str ='xlsx'):
        '''checks path and adds default values if not provided'''

        #print(income_path)
        if re.search(rf'(?:.+\.{re.escape(file_ext)}\Z)', income_path): # path contains filename
            save_path = os.path.normpath(income_path)
            #print('1')
        else:
                
            if re.search(rf'(?<=\\|/)(?:.+\.[A-Za-z]+\Z)', income_path): # check filename with extension
                save_path = os.path.normpath(re.sub(rf'(?<=\\|/)(?:[^\\/]+\.[A-Za-z]+\Z)', filename, income_path))
                print(save_path)
                #print('2')
            else:
                #print('3')
                if re.search(r'(?:\\|/)\Z', income_path): # check '/' at the end of path
                    save_path = os.path.normpath(rf'{income_path}{filename}')
                    #print('31')
                else:
                    save_path = os.path.normpath(rf'{income_path}/{filename}')
                    #print('32')
        
        if re.match(r'[A-Z]\:', save_path) != None or save_path[0] == '.': # path contains letters
            pass
        else:
            save_path = re.sub(r'\A[\\/]*(?=\w)', r'.\\', save_path)
        
        return save_path
        
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###

    def _excel_dates_control(self, series, year=1900, month=1, day=1, hour=None, minute=None):
        '''Excel can't handle the time before 01.01.1900, this func fixes it'''
        
        if series.dtype == '<M8[ns]':
            return series.where(series > pd.Timestamp('01.01.1900 0:00:00'), pd.to_datetime(
                {'year': series.dt.year if year is None else year,
                 'month': series.dt.month if month is None else month,
                 'day': series.dt.day if day is None else day,
                 'hour': series.dt.hour if hour is None else hour,
                 'minute': series.dt.minute if minute is None else minute}))
        else:
            return series
    

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    def _merged_test(self, cell, merged_list):
        '''checks if the cell is merged'''
        
        if isinstance(cell, openpyxl.worksheet.cell_range.CellRange):
            for merged_cell in merged_list:
                try:
                    cell.intersection(merged_cell)
                    return True
                except:
                    pass
        else:
            for merged_cell in merged_list:
                
                if cell.coordinate in merged_cell:
                    return True
        return False
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _get_merged_range(self, cell, merged_list):
        '''returns the range of merged cells if the given cell is part of this range'''
        
        if isinstance(cell, openpyxl.worksheet.cell_range.CellRange):
            result = []
            for merged_cell in merged_list:
                try:
                    cell.intersection(merged_cell)
                    result.append(merged_cell)
                except:
                    pass
            if result:
                return result
        
        else:
            for merged_cell in merged_list:
                if cell.coordinate in merged_cell:
                    return merged_cell
        return cell.coordinate
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _get_fdigitidx(self, coord, alphabet_dict=None):
        '''detects a place of the first digit in the cell coordinate'''
        
        if not alphabet_dict:
            alphabet = list(string.ascii_uppercase)
            excel_alphabet = alphabet + [f'{char}{char2}' for char in alphabet for char2 in alphabet]
            alphabet_dict = {key: value for value, key in enumerate(excel_alphabet, start=1)}
            
        coord_list = list(coord)
        for elem in coord_list:
            if not elem in alphabet_dict:
                return coord_list.index(elem)

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _change_coord(self, coord, row=None, col=None, alphabet_dict=None):
        '''returns a new coordinate ("А1" format) shifted by given cell on given number of rows and columns'''
        
        if not alphabet_dict:
            alphabet = list(string.ascii_uppercase)
            excel_alphabet = alphabet + [f'{char}{char2}' for char in alphabet for char2 in alphabet]
            alphabet_dict = {key: value for value, key in enumerate(excel_alphabet, start=1)}
        
        coord = str(coord)
        
        fdigitidx = self._get_fdigitidx(coord)
        
        old_row = ''.join(list(coord)[fdigitidx:])
        if row:
            new_row = str(int(old_row) + row)
        else:
            new_row = old_row
        
        old_col = ''.join(list(coord)[:fdigitidx])
        if col:
            new_col = excel_alphabet[alphabet_dict[old_col] + col - 1]
        else:
            new_col = old_col
           
        return new_col + new_row
        
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _copy_style(self, source_cell, target_cell):
    
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###

    def _insert_rows(self, cell_coord, rows_number, copymerge=True):
        '''inserts rows correctly managing merged range and cell sizes''' 
        
        row_num = self.sheet[cell_coord].row
        
        # saving original row heights
        orig_rowheights = {i: row_h.height for i, row_h in self.sheet.row_dimensions.items()}
        #print(orig_rowheights)
        mergetocopy_list = []
        
        for m_range in self.sheet.merged_cells.ranges:
            if row_num <= m_range.max_row:
                if row_num > m_range.min_row:
                    m_range.expand(down=rows_number)
                else:
                    m_range.shift(row_shift=rows_number)
            elif row_num == m_range.max_row + 1:
                mergetocopy_list.append([m_range.min_col, m_range.max_col])
        
        self.sheet.insert_rows(row_num, rows_number)
        
        for row in range(rows_number):
            row_letter = row_num + row
            
            # copy row heights from the previous row
            try:
                self.sheet.row_dimensions[row_letter].height = orig_rowheights[row_num - 1]
            except:
                pass
    
            # copying styles
            for cell in self.sheet[row_letter]:
                celltocopy = self.sheet[openpyxl.utils.cell.coordinate_from_string(cell.coordinate)[0] + str(row_num - 1)]
                
                self._copy_style(celltocopy, cell)
                            
            # merge cells as it is in the previous row
            if copymerge:
                for col_range in mergetocopy_list:
                    self.sheet.merge_cells(start_column=col_range[0], start_row=row_letter, end_column=col_range[1], end_row=row_letter)
    
        for row_index in range(row_letter+1, self.sheet.max_row):
            try:
                self.sheet.row_dimensions[row_index].height = orig_rowheights[row_num]
            except:
                pass
            row_num += 1
        self.sheet.merged_cells.ranges = self.sheet.merged_cells.ranges
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _insert_columns(self, cell_coord, cols_number, copymerge=True):
        '''inserts columns correctly managing merged range and cell sizes'''
        col_num = self.sheet[cell_coord].column
        #a = self.sheet[cell_coord]
        #print('col_num:', col_num)
        #print('Cell to write1:', a.coordinate)
         # saving original col heights
        orig_colwidths = {i: col_w.width for i, col_w in self.sheet.column_dimensions.items()}
        #print('orig_colwidths:', orig_colwidths)
        
        mergetocopy_list = []
        for m_range in self.sheet.merged_cells.ranges:
            if col_num <= m_range.max_col:
                if col_num > m_range.min_col:
                    m_range.expand(right=cols_number)
                else:
                    m_range.shift(col_shift=cols_number)
                    
            elif col_num == m_range.max_col + 1:
                mergetocopy_list.append([m_range.min_row, m_range.max_row])
        #print('Cell to write2:', a.coordinate)            
        self.sheet.insert_cols(col_num, cols_number)
        
        #print('Cell to write3:', a.coordinate)
    
        coltocopy = openpyxl.utils.cell.get_column_letter(col_num - 1)
        for col in range(cols_number):
    
            col_letter = openpyxl.utils.cell.get_column_letter(col_num + col)
    
            # copy row heights from the previous row
            try:
                self.sheet.column_dimensions[col_letter].width = orig_colwidths[coltocopy]
            except:
                pass
    
            # copying styles
            for cell in self.sheet[col_letter]:
                celltocopy = self.sheet[coltocopy + str(cell.row)]
                self._copy_style(celltocopy, cell)
                
            # merge cells as it is in the previous row
            if copymerge:
                for row_range in mergetocopy_list:
                    self.sheet.merge_cells(start_column=col_num+col, start_row=row_range[0], end_column=col_num+col, end_row=row_range[1])
    
        for col_index in range(col_num+col+1, self.sheet.max_column):
            try:
                self.sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(col_index)].width = orig_colwidths[openpyxl.utils.cell.get_column_letter(col_num)]
            except:
                pass
            col_num += 1
            
        self.sheet.merged_cells.ranges = self.sheet.merged_cells.ranges

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _copy_merge(self, source_cell, target_cell, debug=False):
    
        mc_ranges = self.sheet.merged_cells
        
        if isinstance(source_cell, openpyxl.worksheet.cell_range.CellRange) or self._merged_test(source_cell, mc_ranges) and source_cell != target_cell:
            logging.debug('Starting merge copy')
            if isinstance(source_cell, openpyxl.worksheet.cell_range.CellRange):
                source_range = source_cell

            else:
            
                source_range = self._get_merged_range(source_cell, mc_ranges)
                if debug:
                    logging.debug(f'source_range: {source_range}')
            source_size = source_range.size
            if debug:
                logging.debug(f'source_size: {source_size}')
            target_range = f'{target_cell.coordinate}:{self._change_coord(target_cell.coordinate, row=source_size["rows"]-1, col=source_size["columns"]-1)}'
            if debug:
                logging.debug(f'target_range: {target_range}')
            target_range_rangetype = openpyxl.worksheet.cell_range.CellRange(target_range)
            #if target_cell.column == 1:
            #    print(sheet.merged_cells.ranges)

            if self._merged_test(target_range_rangetype, mc_ranges):
                
                
                intersection_list = self._get_merged_range(target_range_rangetype, mc_ranges)
                self.sheet.merged_cells.ranges = self.sheet.merged_cells.ranges
                for itrs in intersection_list:
                    
                    self.sheet.unmerge_cells(itrs.coord)
                    
            
                 
            self.sheet.merge_cells(target_range)
            if debug:
                logging.debug(f'merged {target_range}')
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _get_colname_list(self, col_range_start, col_range_end, col_num_list, col_name_list, table_num, cell):
    
        if col_range_start:
            result = list(self.income_data[table_num].columns[int(col_range_start[2])-1:int(col_range_end[2])])
            
        elif col_num_list:
            result = list(self.income_data[table_num].columns[col_num_list])
                                        
        elif col_name_list:
            result = col_name_list
                                        
        else:
            result = []
            if not 'wrong template syntax' in self.errors:
                self.errors['wrong template syntax'] = []
            self.errors['wrong template syntax'].append(cell.coordinate)
        return result

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    def _write_error(self, cell, array, sheet, orientation='row', error_mes='source or template error'):
        cell.value = f'{error_mes}: {array}'
        cell.font = openpyxl.styles.Font(color='00FFFFFF')
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="00FF0000")
        if orientation == 'column':
            end_cell = sheet[self._change_coord(cell.coordinate, row=1)]
            
        else:
            end_cell = sheet[self._change_coord(cell.coordinate, col=1)]
            
        return end_cell

    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###

    def check_step(self, cell, step, orient='row'):

        startcell_size = None
        step = int(step)
        merge_result = self._get_merged_range(cell, self.sheet.merged_cells)
        if isinstance(merge_result, openpyxl.worksheet.cell_range.CellRange):
            startcell_size = merge_result.size

        if startcell_size:
            if orient == 'row':
                new_step = startcell_size['rows']
            else:
                new_step = startcell_size['columns']

            
            if new_step > step:
                    
                #if not 'wrong step in the template' in self.errors:
                #    self.errors['wrong step in the template'] = []
                #self.errors['wrong step in the template'].append({cell.coordinate: f'start cell size: {startcell_size}, step in the template: {step}'})
                step = new_step
        return step
                
        
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###

    
    def _write_array(self, array, orientation, merge, write_mode, start_cell, step=1, step_check=False, debug=False):
        
        ''' Writes an array (1D) from start_cell in the direction depending on the given orientation.
            Merges the cells with equal values if 'merge' is True.
        '''
        if debug:
            logging.debug('Starting to write the array')
        
        cell_to_write = start_cell
        cell_to_check = start_cell
        
        step = int(step)
        if step_check:
            step = self.check_step(start_cell, step, orientation)
      
        if orientation == 'row':
            i = 0
                            
            for value in array:
                i += 1
                if debug:
                    
                    logging.debug(f'step: {i}')
                    logging.debug(f'cell_to_write: {cell_to_write.coordinate}')

                # copying merge
                
                #print('After copy merge', self.sheet._cells)
                #print('After copy merge', self.sheet.merged_cells)
                #print('ctr', cell_to_write)
                #print('value',  value)

               
                
                cell_to_write.value = value
                
                 # copying styles
                self._copy_style(start_cell, cell_to_write)
                self._copy_merge(start_cell, cell_to_write, debug=debug)
                
                
    
                # copying cell size
                try:
                    self.sheet.row_dimensions[cell_to_write.row].height = self.sheet.row_dimensions[start_cell.row].height
                except:
                    pass
                
                cell_to_write = self.sheet[self._change_coord(cell_to_write.coordinate, row=step)]
                    
                # insert rows if specified
                if write_mode == 'insert' and i < len(array):
                    self._insert_rows(cell_to_write.coordinate, step, copymerge=False)
                    cell_to_write = self.sheet[self._change_coord(cell_to_write.coordinate, row=-step)]
                                
            # merge equal adjacent cells if specified
            if merge:
                
                merge_range = ''
                range_start = cell_to_check.coordinate
                
                for value in array:
                    
                    range_end = self.sheet[self._change_coord(cell_to_check.coordinate, row=step)].coordinate
                    
                    if self.sheet[range_end].value == value:
                        if self._merged_test(self.sheet[range_start], self.sheet.merged_cells):
                            mer_range = self._get_merged_range(self.sheet[range_start], self.sheet.merged_cells)
                            self.sheet.unmerge_cells(mer_range.coord)
                                            
                        if self._merged_test(self.sheet[range_end], self.sheet.merged_cells):
                            mer_range = self._get_merged_range(self.sheet[range_end], self.sheet.merged_cells)
                            range_end = self.sheet.cell(row=mer_range.max_row,
                                                   column=self.sheet[range_end].column).coordinate
                            self.sheet.unmerge_cells(mer_range.coord)
                            
                        merge_range = f'{range_start}:{range_end}'
                        
                        cell_to_check = self.sheet[self._change_coord(cell_to_check.coordinate, row=step)]
                        
                                                    
                    else:
                        if merge_range:
                        # проверка не попадает ли объединяемый диапазон на существующие объединенные диапазоны
                            self.sheet.merge_cells(merge_range)
                            merge_range = ''
                        range_start = range_end
                        cell_to_check = self.sheet[self._change_coord(cell_to_check.coordinate, row=step)]
       
            
        else:
            i = 0
                            
            for value in array:
                i += 1
                #print('Step:', i)
                
                
                cell_to_write.value = value
                self._copy_style(start_cell, cell_to_write)
                self._copy_merge(start_cell, cell_to_write, debug=debug)
                
                try:
                    self.sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(cell_to_write.column)].width = self.sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(start_cell.column)].width
                except:
                    pass

                cell_to_write = self.sheet[self._change_coord(cell_to_write.coordinate, col=step)]

                if write_mode == 'insert' and i < len(array):
                    
                    self._insert_columns(cell_to_write.coordinate, step, copymerge=False)
                    cell_to_write = self.sheet[self._change_coord(cell_to_write.coordinate, col=-step)]
                    
    
            if merge:
               
                merge_range = ''
                range_start = cell_to_check.coordinate
                for value in array:
                    
                    range_end = self.sheet[self._change_coord(cell_to_check.coordinate, col=step)].coordinate
                                    
                    if self.sheet[range_end].value == value:
                        if self._merged_test(self.sheet[range_start], self.sheet.merged_cells):
                            mer_range = self._get_merged_range(self.sheet[range_start], self.sheet.merged_cells)
                            self.sheet.unmerge_cells(mer_range.coord)
                                            
                        if self._merged_test(self.sheet[range_end], self.sheet.merged_cells):
                            mer_range = self._get_merged_range(self.sheet[range_end], self.sheet.merged_cells)
                            range_end = self.sheet.cell(row=self.sheet[range_end].row,
                                                   column=mer_range.max_col).coordinate
                            self.sheet.unmerge_cells(mer_range.coord)
                            
                        
                        merge_range = f'{range_start}:{range_end}'
                        cell_to_check = self.sheet[self._change_coord(cell_to_check.coordinate, col=step)]
                                                
                    else:
                        if merge_range:
                        # проверка не попадает ли объединяемый диапазон на существующие объединенные диапазоны
                            self.sheet.merge_cells(merge_range)
                            #print('Merged!!!', merge_range)
                            merge_range = ''
                        range_start = range_end
                        cell_to_check = self.sheet[self._change_coord(cell_to_check.coordinate, col=step)]
                    #print('merge range:', merge_range)
                    #print('cell_to_check in the end:', cell_to_check)
                
        end_cell = cell_to_write
        return end_cell
    
    ###||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||###
    
    
    
    def __init__(self,
                 template_path: str | None = None,
                 report_path: str | None = None,
                 log_path: str | None = None):
        
        if template_path:
            self.template_path = self._path_check(template_path, 'template.xlsx')
        else:
            self.template_path = './template.xlsx'
            
        current_datetime = pd.Timestamp.now()
        default_filename = f'report_{current_datetime.strftime("%d-%m-%y_%H-%M-%S-%f")}.xlsx'
        
        if report_path:
            self.report_path = self._path_check(report_path, default_filename)
        else:
            self.report_path = f'./{default_filename}'

        report_filename = re.findall(r"[^\\/]+(?=\.xlsx)", self.report_path)[0]
        default_logname = f'log_{report_filename}.log'
        if log_path:
            if log_path == 'no log':
                self.log_path = None
            else:
                self.log_path = self._path_check(log_path, default_logname)
        else:
            self.log_path = f'./{default_logname}'
        

    def write(self,
              variables: pd.Series | dict | None = None,
              tables: pd.DataFrame | pd.Series | list | tuple | None = None):

        if self.log_path:
            logging.basicConfig(level=logging.DEBUG, filename=self.log_path, filemode="w", format="%(relativeCreated)d %(lineno)d %(asctime)s %(levelname)s %(message)s", force=True)
        else:
            logging.basicConfig(level=logging.CRITICAL, filename=self.log_path, filemode="w", format="%(relativeCreated)d %(lineno)d %(asctime)s %(levelname)s %(message)s", force=True)
        
        if isinstance(variables, type(None)) and isinstance(tables, type(None)):
            self.errors['nothing to write'] = True
            
        else:
            if not isinstance(variables, type(None)):
                self.variables = variables
            else:
                self.variables = pd.Series()

            
            if not isinstance(tables, type(None)):
                if isinstance(tables, pd.DataFrame):
                    
                    self.income_data = {1: tables.apply(self._excel_dates_control)}
                    
                else:
                    self.income_data = {i+1: data.apply(self._excel_dates_control) for i, data in enumerate([table for table in tables if isinstance(table, pd.DataFrame)])}
            
            ##### loading template
            print(f'Processing template {self.template_path}...')
            try:
                self.wb = load_workbook(self.template_path)
                
            except:
                print('!!! wrong template path or no template file !!!')
                sys.exit()

            #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
            ##### starting template scanning and writing data
            #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
            
            print(f'{len(self.wb.sheetnames)} sheets found')
            for sheet_name in self.wb.sheetnames:
                self.writing_history[sheet_name] = pd.DataFrame()
                self.sheet = self.wb[sheet_name]
                logging.debug(self.wb)
                logging.debug(f'Processing sheet: {self.sheet}')
                print(f'Processing sheet: {sheet_name}')
                
                merged_list = self.sheet.merged_cells
                
                logging.debug('Analysing template for pivot table')
                print('Analysing for pivot table...')
                # scaning pivot tables
                pivot_tables = {}
                
                
                for row in self.sheet:
                    for cell in row:
                        if cell.value:
                            array = re.findall(r'(?<=\[\[).+(?=\]\])', str(cell.value))
                            if array:
                                logging.debug(f'cell: {cell.coordinate}')
                                
                                pivot_table = re.search(r'(?<=\Apt)(\d+)', array[0])
                                if pivot_table:
                                    pivot_table = int(pivot_table[0])
                                    if not pivot_table in pivot_tables:
                                        pivot_tables[pivot_table] = {}
                                    #print('pivot_table:', pivot_table)
                                    logging.debug(f'pivot_table: {pivot_table}')
                                    col_name = re.search(r'(?<=\.)(\s*)([A-Za-z_\-0-9]+)', array[0])
                                    col_range_start = re.search(r'(?<=\()(\s*)(\d+)(\s*)(?=:)', array[0])
                                    col_range_end = re.search(r'(?<=:)(\s*)(\d+)(\s*)(?=\))', array[0])
                                    
                                    col_list = re.search(r'(?<=\()(\s*)([A-Za-z_0-9 ,]+)(\s*)(?=\))', array[0])
                                    col_num_list = None
                                    col_name_list = None
                                    
                                    
                                    if col_list:
                                        col_list = re.split(r'\s*,\s*', col_list[2])
                                        try:
                                            col_num_list = [(int(el)-1) for el in col_list]
                                        except:
                                            col_name_list = col_list
                                         
                                    array_params = re.search(r'(?<=/)(\s*)([a-z*0-9 ]+\Z)', array[0])
                                    
                                    if array_params:
                                        array_params = array_params[2]
                                    else:
                                        array_params = ''
                                    
                                    if 'c' in array_params:
                                        orientation = 'column'
                                            
                                    elif 'v' in array_params:
                                        orientation = 'values'
                                            
                                    else:
                                        orientation = 'row'
                                        
                                    if not orientation in pivot_tables[pivot_table]:
                                        pivot_tables[pivot_table][orientation] = []
                                    
                                    try:
                                        source_data = self.income_data[pivot_table]
                                        #print(source_data)
                                        logging.debug(f'found the pivot_table: {pivot_table}')
                                        
                                    except:
                                        source_data = pd.DataFrame()
                                        if not 'can`t find this number of pivot table in income data' in self.errors:
                                            self.errors['can`t find this number of pivot table in income data'] = []
                                        self.errors['can`t find this number of pivot table in income data'].append({cell.coordinate: pivot_table})
                                    
                                    if not source_data.empty:
                                        if col_name:
                                            col_name = col_name[2]
                                            #print('col_name:', col_name)
                                            if col_name in source_data:
                                                pivot_tables[pivot_table][orientation].append(col_name)
                                            else:
                                                if col_name != '__VAL_NAMES__':
                                                    if not 'pivot scan: wrong column name or index' in self.errors:
                                                        self.errors['pivot scan: wrong column name or index'] = []
                                                    self.errors['pivot scan: wrong column name or index'].append({cell.coordinate: f'{col_name} -> pt{pivot_table}'})
                                        
                                        else:
                                            table_num = pivot_table
                                            col_name_list = self._get_colname_list(col_range_start, col_range_end, col_num_list, col_name_list, table_num, cell)  ##########
                                            #print('col_name_list:', col_name_list)
                                            try:
                                                test = source_data[col_name_list]
                                                pivot_tables[pivot_table][orientation].extend(col_name_list)
                                            except:
                                                if not 'wrong column name or index' in self.errors:
                                                    self.errors['wrong column name or index'] = []
                                                self.errors['wrong column name or index'].append({cell.coordinate: f'{col_name_list} -> t{table_num}'})
                                                
                
                #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#                                
                ##### creating dataframes for found pivot tables
                #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
                
                logging.debug(f'Find pivot_tables:\n{pivot_tables}')
                logging.debug('Creating pivot dataframes')
                pivot_dfs = {}
                for pt in pivot_tables:
                    try:
                        pivot_df = pd.pivot(source_data,
                                            index=pivot_tables[pt]['row'],
                                            columns=pivot_tables[pt]['column'],
                                            values=pivot_tables[pt]['values'])
                        pivot_dfs[pt] = pivot_df
                        
                        
                    except:
                        logging.debug(str(['#']*50))
                        logging.error(f'pivot{pt}: creation error', exc_info=True)
                        logging.debug(str(['#']*50))
                        if not f'pivot{pt}: creation error' in self.errors:
                            self.errors[f'pivot{pt}: creation error'] = []
                        self.errors[f'pivot{pt}: creation error'].append(pivot_tables[pt])
                        
                logging.debug(f'Created pivot dataframes: {pivot_dfs}')
                if pivot_dfs:
                    print(f'Successfully found {len(pivot_dfs)} pivot_tables')
                else:
                    print('No pivot tables')

                #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#                                
                ##### starting to write data
                #\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

                print('Writing a report...')
                l_sheet = list(self.sheet)    

                logging.debug('Starting write a report')
                
                for row in l_sheet:
                    
                    logging.debug(f'Processing {row} in l_sheet {l_sheet}')

                    l_row = list(self.sheet[row[0].row])
                    logging.debug(f'l_row {l_row}')
                    
                    for cell in l_row:
                        # check cells for values
                        logging.debug(f'cell: {cell}')
                        if cell.value:
                                                        
                            writing_status = 'Success'

                            #### processing simple variables
                            simple_vars = re.findall(r'(?<=%%)[A-Za-z_\-0-9]+(?=%%)', str(cell.value))
                            if simple_vars:
                                logging.debug(f'simple_vars: {simple_vars}')
                                to_write = cell.value
                                for simple_var in simple_vars:
                                    
                                    try:
                                        to_write = to_write.replace(f'%%{simple_var}%%', str(self.variables[simple_var]))
                                        
                                    except:
                                        to_write = to_write.replace(f'%%{simple_var}%%', '<NO DATA>')
                                        if not 'no income data for simple_vars' in self.errors:
                                            self.errors['no income data for simple_vars'] = []
                                        self.errors['no income data for simple_vars'].append(simple_var)
                                        writing_status = 'source or template error'
                                            
                                cell.value = to_write
                                
                                logging.debug(f'to_write: {to_write}')
                                logging.debug(f'writing_status: {writing_status}')
                                record = pd.DataFrame({'cell': cell.coordinate, 'type': 'simple variable', 'source': str(simple_vars), 'status': writing_status}, index=[0])
                                self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                        record],
                                                                        ignore_index=True)
                            
                            #### processing arrays (tables)            
                            array = re.findall(r'(?<=\[\[).+(?=\]\])', str(cell.value))
                            if array:

                                pivot_table = re.search(r'(?<=\Apt)(\d+)', array[0])
                                table_num = re.search(r'(?<=\At)(\d+)|(?<=pt)(\d+)', array[0])
                                
                                if pivot_table:
                                    pivot_table = int(pivot_table[0])
                                
                                elif table_num:
                                    table_num = int(table_num[0])
                                                                    
                                else:
                                    pivot_table = None
                                    table_num = 1
                                
                                logging.debug(f'pivot_table: {pivot_table}')
                                logging.debug(f'table_num: {table_num}')
                                    
                                col_name = re.search(r'(?<=\.)(\s*)([A-Za-z_\-0-9]+)', array[0])
                                
                                col_range_start = re.search(r'(?<=\()(\s*)(\d+)(\s*)(?=:)', array[0])
                                col_range_end = re.search(r'(?<=:)(\s*)(\d+)(\s*)(?=\))', array[0])
                                
                                col_list = re.search(r'(?<=\()(\s*)([A-Za-z_\-0-9 ,]+)(\s*)(?=\))', array[0])
                                
                                if all([not elem for elem in [col_name, col_range_start, col_list]]):
                                    self._write_error(cell, array, self.sheet)
                                        
                                col_num_list = None
                                col_name_list = None
                                if col_list:
                                    col_list = re.split(r'\s*,\s*', col_list[2])
                                    try:
                                        col_num_list = [int(el)-1 for el in col_list]
                                    except:
                                        col_name_list = col_list
                                
                                
                                #print('col to write', col_name)
                                logging.debug(f'col_name: {col_name}')
                                logging.debug(f'col_range_start: {col_range_start}')
                                logging.debug(f'col_range_end: {col_range_end}')
                                logging.debug(f'col_list: {col_list}')
                                logging.debug(f'col_num_list: {col_num_list}')
                                logging.debug(f'col_name_list: {col_name_list}')
                                
                                array_params = re.search(r'(?<=/)(\s*)([a-z0-9* ]+\Z)', array[0])
                                if array_params:
                                    array_params = array_params[2]
                                else:
                                    array_params = ''
                                    
                                if 'c' in array_params and 'r' in array_params:
                                    orientation = 'row'
                                    if not 'template syntax errors' in self.errors:
                                        self.errors['template syntax errors'] = []
                                    self.errors['template syntax errors'].append({cell.coordinate: 'ambiguous orientation keys'})
                                elif 'c' in array_params:
                                    orientation = 'column'
                                elif 'v' in array_params:
                                    orientation = 'values'
                                else:
                                    orientation = 'row'
                                
                                logging.debug(f'orientation: {orientation}')
                                
                                if 'i' in array_params and 'u' in array_params:
                                    write_mode = 'update'
                                    if not 'template syntax errors' in self.errors:
                                        self.errors['template syntax errors'] = []
                                    self.errors['template syntax errors'].append({cell.coordinate: 'ambiguous write mode keys'})
                                elif 'i' in array_params:
                                    write_mode = 'insert'
                                else:
                                    write_mode = 'update'
                                
                                logging.debug(f'write_mode: {write_mode}')
                                    
                                if '*' in array_params:
                                    merge = True
                                else:
                                    merge = False
                                logging.debug(f'merge: {merge}')
                                
                                
                                step = re.search(r'(?<=s)\d+', array_params)
                                if step:
                                    step = step[0]
                                else:
                                    step = 1

                                ver_step = self.check_step(cell, step, orient='row')
                                hor_step = self.check_step(cell, step, orient='column')
                                
                                logging.debug(f'step: {step}')
                                if pivot_table:
                                    
                                    logging.debug('writing pivot_table')
                                    try:
                                        df_to_write = pivot_dfs[pivot_table]
                                        
                                    except:
                                        df_to_write = pd.DataFrame()
                                        if not 'no corresponding pivot table' in self.errors:
                                            self.errors['no corresponding pivot table'] = []
                                        self.errors['no corresponding pivot table'].append({cell.coordinate: f'pt{pivot_table}'})
                                        logging.debug(str(['#']*50))
                                        logging.error('no corresponding pivot table', exc_info=True)
                                        logging.debug(str(['#']*50))
                                        
                                else:
                                    
                                    logging.debug('writing simple table')
                                    try:
                                        df_to_write = self.income_data[table_num]
                                    
                                    except:
                                        df_to_write = pd.DataFrame()
                                        if not 'can`t find the table number in the income data' in self.errors:
                                            self.errors['can`t find the table number in the income data'] = []
                                        self.errors['can`t find the table number in the income data'].append({cell.coordinate: f'port {table_num}'})
                                        logging.debug(str(['#']*50))
                                        logging.error('can`t find the table number in the income data', exc_info=True)
                                        logging.debug(str(['#']*50))
   
                                if not df_to_write.empty:

                                    ########################################################            
                                    ##### writting pivot table elements ####################
                                    ########################################################
                                    if pivot_table:
                                        
                                            ##################
                                        ####  row orientation ####
                                            ##################
                                        
                                        if orientation == 'row':
                                            
                                            if col_name:
                                                
                                                ###  writing 1D-array ################
                                                col_name = col_name[2]
                                                logging.debug('Writing 1D array')
                                                logging.debug(f'col_name: {col_name}')
                                                
                                                try:
                                                    col_to_write = df_to_write.index.get_level_values(col_name).to_series()
                                                    self._write_array(col_to_write, orientation, merge, write_mode, cell, ver_step)
                                                    # row extension
                                                    l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])
                                                    
                                                except:
                                                    self._write_error(cell, array, self.sheet)
                                                    writing_status = 'source or template error'
                                                    logging.debug(str(['#']*50))
                                                    logging.error('source or template error', exc_info=True)
                                                    logging.debug(str(['#']*50))
                                                    
                                                record = pd.DataFrame({'cell': cell.coordinate, 'type': 'pivot table', 'array_shape': '1D',
                                                                       'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                       'write_step': step, 'source': f'table: {table_num}, col_name: {col_name}',
                                                                      'status': writing_status}, index=[0])

                                                #logging.debug(f'to_write: {col_to_write}')
                                                logging.debug(f'writing_status: {writing_status}')
                                                self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                        record],
                                                                                        ignore_index=True)
                                             
                                            else:
                                                ###  writing 2D-array ################
                                                logging.debug('Writing 2D array')
                                                table_num = int(table_num[0])
                                                col_name_list = self._get_colname_list(col_range_start, col_range_end, col_num_list, col_name_list, table_num, cell)

                                               
                                                                                        
                                                logging.debug(f'col_name_list: {col_name_list}')
                                                cell_to_write = cell
                                                
                                                start_range = self._get_merged_range(cell, self.sheet.merged_cells)
                                                
                                                ## iterating a list of column names
                                                for i, col in enumerate(col_name_list):
                                                    logging.debug(f'###{i}###: {cell_to_write} <= {col}')
                                                     # to transmit information about start cell to writing history
                                                    cell_info = cell_to_write
                                                    try:
                                                        col_to_write = df_to_write.index.get_level_values(col).to_series()
                                                        
                                                        if i == 0:
                                                            end_cell = self._write_array(col_to_write, orientation,
                                                                                       merge, write_mode,
                                                                                       cell_to_write, ver_step)
                                                        else:
                                                            end_cell = self._write_array(col_to_write,
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode='update',
                                                                                       start_cell=cell_to_write,
                                                                                       step=ver_step)
                                                        # row extension
                                                        l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])
                                                                                                            
                                                        if i < (len(col_name_list) - 1):
                                                            cell_to_write = self.sheet[self._change_coord(end_cell.coordinate,
                                                                                                   row=-len(col_to_write)*ver_step,
                                                                                                   col=hor_step)]
                                                            self._copy_style(cell, cell_to_write)
                                                            if isinstance(start_range, openpyxl.worksheet.cell_range.CellRange):
                                                                self._copy_merge(start_range, cell_to_write)
                                                       
                                                    
                                                        #if write_mode == 'insert':
                                                        #    insert_columns(cell_to_write.coordinate, 1, sheet)
                                                        #    cell_to_write = sheet[change_coord(cell_to_write.coordinate, col=-1)]
                                                        
                                                    except:
                                                        cell_to_write = self._write_error(cell_to_write, array, self.sheet, orientation)
                                                        writing_status = 'source or template error'
                                                        logging.debug(str(['#']*50))
                                                        logging.error('source or template error', exc_info=True)
                                                        logging.debug(str(['#']*50))
                                                        
                                                    #logging.debug(f'to_write: {col_to_write}')
                                                    logging.debug(f'writing_status: {writing_status}')
                                                    
                                                    record = pd.DataFrame({'cell': cell_info.coordinate, 'type': 'pivot table', 'array_shape': '2D',
                                                                           'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                           'write_step': step, 'source': f'table: {table_num}, col_name: {col}',
                                                                          'status': writing_status}, index=[0])
                                                
                                                    self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                            record],
                                                                                            ignore_index=True)
                                                        
                                            ####################
                                        #### column orientation ####
                                            ####################
                                        
                                        elif orientation == 'column':
                                            if col_name:
                                                
                                                ###  writing 1D-array ################
                                                logging.debug('Writing 1D array')
                                                logging.debug(f'col_name: {col_name}')
                                                try:
                                                    col_name = col_name[2]
                                                    if col_name == '__VAL_NAMES__':
                                                        col_to_write = df_to_write.columns.get_level_values(0).to_series()
                                                    else:
                                                        col_to_write = df_to_write.columns.get_level_values(col_name).to_series()

                                                    
                                                    
                                                    self._write_array(col_to_write, orientation, merge, write_mode, cell, hor_step)
                                                    #print('l_row', l_row)
                                                    if write_mode == 'insert':
                                                        l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])
                                                    #print('col_name:', col_name)
                                                
                                                except:
                                                    self._write_error(cell, array, self.sheet)
                                                    writing_status = 'source or template error'
                                                    logging.debug(str(['#']*50))
                                                    logging.error('source or template error', exc_info=True)
                                                    logging.debug(str(['#']*50))

                                                #logging.debug(f'to_write: {col_to_write}')
                                                logging.debug(f'writing_status: {writing_status}')
                                                    
                                                record = pd.DataFrame({'cell': cell.coordinate, 'type': 'pivot table', 'array_shape': '1D',
                                                                       'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                       'write_step': step, 'source': f'table: {table_num}, col_name: {col_name}',
                                                                      'status': writing_status}, index=[0])
                                                
                                                self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                        record],
                                                                                        ignore_index=True)
                                                                                 
                                            else:
                                                
                                                ###  writing 2D-array ################
                                                logging.debug('Writing 2D array')
                                                table_num = int(table_num[0])
                                                col_name_list = self._get_colname_list(col_range_start, col_range_end, col_num_list, col_name_list, table_num, cell)

                                                logging.debug(f'col_name_list: {col_name_list}')
                                                cell_to_write = cell

                                                start_range = self._get_merged_range(cell, self.sheet.merged_cells)
                                                ## iterating a list of column names
                                                for i, col in enumerate(col_name_list):
                                                    logging.debug(f'###{i}###: {cell_to_write} <= {col}')
                                                     # to transmit information about start cell to writing history
                                                    cell_info = cell_to_write
                                                    try:
                                                        col_to_write = df_to_write.columns.get_level_values(col).to_series()
                                                        
                                                        if i == 0:
                                                            end_cell = self._write_array(col_to_write,
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode,
                                                                                       cell_to_write,
                                                                                       hor_step)
                                                        else:
                                                            end_cell = self._write_array(col_to_write,
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode='update',
                                                                                       start_cell=cell_to_write,
                                                                                       step=hor_step)
            
                                                        #l_row.extend(sheet[cell.row][sheet[cell.row].index(row[-1])+1:])
                                                    
                                                        
                                                        if i < (len(col_name_list) - 1):
                                                            cell_to_write = self.sheet[self._change_coord(end_cell.coordinate,
                                                                                                   row=ver_step,
                                                                                                   col=-len(col_to_write)*hor_step)]
                                                            self._copy_style(cell, cell_to_write)
                                                            if isinstance(start_range, openpyxl.worksheet.cell_range.CellRange):
                                                                self._copy_merge(start_range, cell_to_write)
                                                    
                                                        #if write_mode == 'insert':
                                                        #    insert_rows(cell_to_write.coordinate, 1, sheet)
                                                        #    cell_to_write = sheet[change_coord(cell_to_write.coordinate, row=-1)]
                                                    except:
                                                        cell_to_write = self._write_error(cell_to_write, array, self.sheet, orientation)
                                                        writing_status = 'source or template error'
                                                        logging.debug(str(['#']*50))
                                                        logging.error('source or template error', exc_info=True)
                                                        logging.debug(str(['#']*50))

                                                    #logging.debug(f'to_write: {col_to_write}')
                                                    logging.debug(f'writing_status: {writing_status}')
                                                    
                                                    record = pd.DataFrame({'cell': cell_info.coordinate, 'type': 'pivot table', 'array_shape': '2D',
                                                                           'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                           'write_step': step, 'source': f'table: {table_num}, col_name: {col}',
                                                                          'status': writing_status}, index=[0])
                                                
                                                    self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                            record],
                                                                                            ignore_index=True)
                                        
                                        
                                            ####################
                                        #### values orientation ####
                                            ####################
                                        
                                        else:
                                            logging.debug('Writing 2D array')
                                            orientation = 'row'
                                            table_num = int(table_num[0])

                                            logging.debug(f'col_name_list: {df_to_write.columns}')
                                        
                                            cell_to_write = cell
                                            hor_step = self.check_step(cell, step, orient='column')
                                            #print('hor_step', hor_step)
                                            for i, col in enumerate(df_to_write):
                                                logging.debug(f'###{i}###: {cell_to_write} <= {col}')
                                                 # to transmit information about start cell to writing history
                                                cell_info = cell_to_write
                                                try:
                                                    if i == 0:
                                                        end_cell = self._write_array(df_to_write[col], orientation,
                                                                                   merge, write_mode, cell_to_write, ver_step)
                                                    else:
                                                        end_cell = self._write_array(df_to_write[col], orientation,
                                                                                   merge, write_mode='update',
                                                                                   start_cell=cell_to_write, step=ver_step)
                                                    #l_row.extend(sheet[cell.row][sheet[cell.row].index(row[-1])+1:])
                                                        #print('end cell:', end_cell)

                                                    
                                                    if i < (len(df_to_write.columns) - 1):
                                                        cell_to_write = self.sheet[self._change_coord(end_cell.coordinate, row=-len(df_to_write[col])*ver_step, col=hor_step)]
                                                        
                                                        self._copy_style(cell, cell_to_write)
                                                        self._copy_merge(cell, cell_to_write)
                                                    
                                                        #if write_mode == 'insert':
                                                            #insert_columns(cell_to_write.coordinate, 1, sheet)
                                                            #cell_to_write = sheet[change_coord(cell_to_write.coordinate, col=-1)]
                                                            
                                                except:
                                                    cell_to_write = self._write_error(cell_to_write, array, self.sheet, orientation)
                                                    writing_status = 'source or template error'
                                                    logging.debug(str(['#']*50))
                                                    logging.error('source or template error', exc_info=True)
                                                    logging.debug(str(['#']*50))

                                                #logging.debug(f'to_write: {col_to_write}')
                                                logging.debug(f'writing_status: {writing_status}')
                                                    
                                                record = pd.DataFrame({'cell': cell_info.coordinate, 'type': 'pivot table', 'array_shape': '2D',
                                                                           'orientation': 'values', 'write_mode': write_mode, 'merge': merge,
                                                                           'write_step': step, 'source': f'table: {table_num}, col_name: {col}',
                                                                          'status': writing_status}, index=[0])
                                                
                                                self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                            record],
                                                                                            ignore_index=True)
                                                        
                                    
                                    #############################################################            
                                    ##### writting 1 column of simple table #####################
                                    #############################################################
                                            
                                    elif col_name:
                                        col_name = col_name[2]
                                        logging.debug(f'writting 1 column: {col_name} (1D array)')
                                        
                                        try:
                                            col_to_write = df_to_write[col_name]
                                            
                                        except:
                                            
                                            col_to_write = pd.Series()
                                            if not 'wrong column name or index' in self.errors:
                                                self.errors['wrong column name or index'] = []
                                            self.errors['wrong column name or index'].append({cell.coordinate: f'{col_name} -> t{table_num}'})
                                            logging.debug(str(['#']*50))
                                            logging.error('wrong column name or index', exc_info=True)
                                            logging.debug(str(['#']*50))
                                            
                                        
                                        if not col_to_write.empty:
                                            
                                            logging.debug(f'l_row: {l_row}')
                                            
                                            
                                            self._write_array(col_to_write, orientation, merge,
                                                            write_mode, cell, step, step_check=True, debug=True)
                                            #print(l_row)
                                            #print(row)
                                            #print(cell)
                                            try:
                                                l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])
                                            except:
                                                pass
                                            #[<Cell 'Лист1'.A6>, <Cell 'Лист1'.B6>]
                                            #(<Cell 'Лист1'.A6>, <MergedCell 'Лист1'.B6>)
                                            
                                            
                                        else:
                                            self._write_error(cell, array, self.sheet)
                                            writing_status = 'source or template error'

                                        #logging.debug(f'to_write: {col_to_write}')
                                        logging.debug(f'writing_status: {writing_status}')
                                                    
                                        record = pd.DataFrame({'cell': cell.coordinate, 'type': 'simple table', 'array_shape': '1D',
                                                                'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                'write_step': step, 'source': f'table: {table_num}, col_name: {col_name}',
                                                                'status': writing_status}, index=[0])
                                                
                                        self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                record],
                                                                                ignore_index=True)
                                            
                               
                                    #############################################################            
                                    ##### writting a simple table (2D array) ####################
                                    #############################################################
                                    
                                    else:
                                        logging.debug('Writing 2D array')
                                        col_name_list = self._get_colname_list(col_range_start, col_range_end, col_num_list, col_name_list, table_num, cell)
                                        logging.debug(f'col_name_list: {col_name_list}')
                                        try:
                                            df_to_write = df_to_write[col_name_list]
                                            
                                        except:
                                            #df_to_write = pd.DataFrame()
                                            if not 'wrong column name or index' in self.errors:
                                                self.errors['wrong column name or index'] = []
                                            self.errors['wrong column name or index'].append({cell.coordinate: f'{col_name} -> t{table_num}'})
                                            logging.debug(str(['#']*50))
                                            logging.error('wrong column name or index', exc_info=True)
                                            logging.debug(str(['#']*50))
                                     
                                        if not df_to_write.empty:
                                            
                                            cell_to_write = cell
                                            start_range = self._get_merged_range(cell, self.sheet.merged_cells)
                                            for i, col in enumerate(col_name_list):
                                                logging.debug(f'###{i}###: {cell_to_write} <= {col}')
                                                # to transmit information about start cell to writing history
                                                cell_info = cell_to_write
                                                
                                                ### row orientation
                                                
                                                try:
                                                    if orientation == 'row':
                                                    
                                                        if i == 0:
                                                            end_cell = self._write_array(df_to_write[col],
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode,
                                                                                       cell_to_write,
                                                                                       ver_step)
                                                        else:
                                                            end_cell = self._write_array(df_to_write[col],
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode='update',
                                                                                       start_cell=cell_to_write,
                                                                                       step=ver_step)
            
                                                        l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])      #
                                                        
                                                        if i < (len(df_to_write.columns) - 1):
                                                            cell_to_write = self.sheet[self._change_coord(end_cell.coordinate,
                                                                                                   row=-len(df_to_write[col])*ver_step,
                                                                                                   col=hor_step)]
                                                            self._copy_style(cell, cell_to_write)
                                                            if isinstance(start_range, openpyxl.worksheet.cell_range.CellRange):
                                                                self._copy_merge(start_range, cell_to_write)
                                                    
                                                    
                                                            #if write_mode == 'insert':
                                                            #    insert_columns(cell_to_write.coordinate, 1, sheet)
                                                            #    cell_to_write = sheet[change_coord(cell_to_write.coordinate, col=-1)]
                                                    
                                                        
                                                    ### col orientation
                                                    else:
                                                        if i == 0:
                                                            end_cell = self._write_array(df_to_write[col],
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode,
                                                                                       cell_to_write,
                                                                                       hor_step)
                                                        else:
                                                            end_cell = self._write_array(df_to_write[col],
                                                                                       orientation,
                                                                                       merge,
                                                                                       write_mode='update',
                                                                                       start_cell=cell_to_write,
                                                                                       step=hor_step)
                                                        
                                                        l_row.extend(self.sheet[cell.row][self.sheet[cell.row].index(row[-1])+1:])       #
                                                    
                                                        if i < (len(df_to_write.columns) - 1):
                                                            cell_to_write = self.sheet[self._change_coord(end_cell.coordinate,
                                                                                                   row=ver_step,
                                                                                                   col=-len(df_to_write[col])*hor_step)]
                                                            self._copy_style(cell, cell_to_write)
                                                            if isinstance(start_range, openpyxl.worksheet.cell_range.CellRange):
                                                                self._copy_merge(start_range, cell_to_write)
                                                    
                                                            #if write_mode == 'insert':
                                                            #    insert_rows(cell_to_write.coordinate, 1, sheet)
                                                            #    cell_to_write = sheet[change_coord(cell_to_write.coordinate, row=-1)]
                                                            #sheet.calculate_dimension()
                                                except:
                                                    cell_to_write = self._write_error(cell_to_write, array, self.sheet, orientation)
                                                    writing_status = 'source or template error'
                                                    logging.debug(str(['#']*50))
                                                    logging.error('source or template error', exc_info=True)
                                                    logging.debug(str(['#']*50))

                                                logging.debug(f'to_write: {df_to_write[col]}')
                                                logging.debug(f'writing_status: {writing_status}')
                                                    
                                                record = pd.DataFrame({'cell': cell_info.coordinate, 'type': 'simple table', 'array_shape': '2D',
                                                                           'orientation': orientation, 'write_mode': write_mode, 'merge': merge,
                                                                           'write_step': step, 'source': f'table: {table_num}, col_name: {col}',
                                                                          'status': writing_status}, index=[0])
                                                
                                                self.writing_history[sheet_name] = pd.concat([self.writing_history[sheet_name],
                                                                                            record],
                                                                                            ignore_index=True)
                                    

                                    logging.debug(f'##################################\nmerged_cells: {self.sheet.merged_cells.ranges}\n##############################################')
                                    #save_path = f'./{str(cell.coordinate)}_{pd.Timestamp.now().strftime("%d-%m-%y_%H-%M-%S-%f")}.xlsx'
                                
                                    #self.wb.save(save_path)
                                else:
                                    self._write_error(cell, array, self.sheet, orientation='row', error_mes='dataframe-to-write is empty')
                
                        else:
                            logging.debug('The cell doesn`t have value')
                print('Finished')

            try:
    
                self.wb.save(self.report_path)
    
            except:
    
                os.makedirs(os.path.split(self.report_path)[0])
                self.wb.save(self.report_path)

            if self.errors:
                self.status = f'errors: {self.errors}'
            else:
                self.status = 'Success'
            print('--------------------------------------------------------------------------------------------------------------------------')
            print('Status:', self.status)
            print('--------------------------------------------------------------------------------------------------------------------------')
            print('Template_path:', self.template_path)
            print('Report_path:', self.report_path)
            print('Log_path:', self.log_path)
            print('See `self.writing_history` for details')